# import sys
# import os
# sys.path = list(set(sys.path + [os.path.join(os.environ['CGPIPELINE'])]))
import re
from PySide2.QtWidgets import *
# from PySide2.QtCore import Qt
from . import treeExcelDataWidget
from openpyxl import load_workbook
from UI.Dialogs import wdg_utils

presetsDict = {"epi_0100": r'^([a-z]{2,5}|[A-Z]{2,5})\W?\w{2,5}\W?(\w{0,4})$',
               "254_0100": r'^(\d{2,4})((_\w{2,5})|([a-zA-Z]{1,5}\w{1,5}))_?\w{0,4}$'}


class SheetDataWidget(QWidget):
    def __init__(self, parent, sheet):
        super(SheetDataWidget, self).__init__(parent)
        # self.newProject = newProject
        self.sheet = sheet
        self.treeData = []

        self.lay_main = QVBoxLayout()
        self.setLayout(self.lay_main)
# ======================== WIDGETS =============================
        # - title row number
        self.titleRowNumber_Lable = QLabel("Title row number:")
        self.titleRowNumber = QSpinBox()
        self.titleRowNumber.setFixedSize(50, 25)
        self.titleRowNumber.setValue(1)
        self.lay_titelRowNumber = wdg_utils.getHorizontalBlockLayout(self.titleRowNumber_Lable, self.titleRowNumber)
        self.lay_titelRowNumber.addStretch()
        # - add/remove buttons
        self.addRowButton = QPushButton('+')
        self.addRowButton.setFixedSize(25, 25)
        self.removeRowButton = QPushButton('-')
        self.removeRowButton.setFixedSize(25, 25)
        self.lay_rowButtons = wdg_utils.getHorizontalBlockLayout(self.addRowButton, self.removeRowButton)
        self.lay_rowButtons.addLayout(self.lay_titelRowNumber)
        self.lay_rowButtons.addStretch()
        # - regexp field
        self.regexpLable = QLabel('regexp:')
        self.regexpField = QLineEdit()
        self.regexpField.setFixedWidth(250)
        # self.regexpField.setAlignment(Qt.AlignLeft)
        # self.regexpField.setText(pattern_01)
        self.regexpPreset_comboBox = QComboBox()
        self.lay_regexp = wdg_utils.getHorizontalBlockLayout(self.regexpLable, self.regexpField,
                                                            self.regexpPreset_comboBox)
        self.lay_rowButtons.addLayout(self.lay_regexp)
        # - list data
        self.listData = QListWidget()
        # - get data button
        self.getDataButton = QPushButton('Get Data')
        self.getDataButton.setFixedSize(150, 40)
        # - test Collect Button
        # self.testCollectButton = QPushButton("Collect")
        # - tree data Widget
        self.treeDataWidget = treeExcelDataWidget.TreeDataWidget(self)

# ======================== CONNECTS =============================
        self.titleRowNumber.valueChanged.connect(self.updateListItem)
        self.addRowButton.clicked.connect(self.addlistItem)
        self.removeRowButton.clicked.connect(self.removeListItem)
        self.getDataButton.clicked.connect(self.getTreeDataButton)
        # self.testCollectButton.clicked.connect(self.collectAllTreeData)
        self.regexpPreset_comboBox.currentIndexChanged.connect(self.setRegexpPreset)
# ======================== LAYOUTS =============================
        self.lay_main.addLayout(self.lay_rowButtons)
        self.lay_main.addWidget(self.listData)
        self.lay_main.addWidget(self.getDataButton)
        self.lay_main.addWidget(self.treeDataWidget)

# =====================================================
        self.addlistItem("name", True)
        self.addlistItem("description", True)
        self.addlistItem("frames_count", True)
        self.setRegexpPresets()

    def addlistItem(self, nameField="", readOnly=False):
        item = QListWidgetItem()
        self.listData.addItem(item)

        itemWdg = ItemListWidget(self, nameField, readOnly)
        titlesData = self.getTitlesData(self.titleRowNumber.value())
        itemWdg.setColumnsList_comboBox(titlesData)

        item.setSizeHint(itemWdg.sizeHint())
        self.listData.setItemWidget(item, itemWdg)

    def updateListItem(self, inpRowNmb):
        if inpRowNmb == 0:
            return
        titlesData = self.getTitlesData(inpRowNmb)
        for idx in range(self.listData.count()):
            item = self.listData.item(idx)
            wdg_combobox = self.listData.itemWidget(item)
            wdg_combobox.setColumnsList_comboBox(titlesData)

    def removeListItem(self):
        itemCount = self.listData.count()
        if itemCount <= 1:
            return
        self.listData.takeItem(itemCount - 1)

    def setRegexpPresets(self):
        presets = [(key, presetsDict[key]) for key in list(presetsDict.keys())]
        for preset in presets:
            self.regexpPreset_comboBox.addItem(*preset)

    def setRegexpPreset(self):
        preset = self.regexpPreset_comboBox.currentData()
        self.regexpField.setText(preset)

    def getTreeDataButton(self):
        userInpTitles = self.getUserTitlesData()
        if not userInpTitles:
            print("name data not found")
            return
        # print(userInpTitles)

        nameField, nameCoord = userInpTitles[0]
        if nameField != 'name':
            print("Unexpected shot name field.")
            return

        cell = self.sheet[nameCoord]
        nameColumn = cell.column
        nameRow = cell.row

        pattern = self.regexpField.text()
        # pattern_01 = r'^([a-z]{2,5}|[A-Z]{2,5})_?\w{2,5}_?(\w{0,4})$'
        # pattern_02 = r'^(\d{2,4})((_\w{2,5})|([a-zA-Z]{1,5}\w{1,5}))_?\w{0,4}$'

        self.treeData = self.collectTreeData(nameColumn, nameRow, userInpTitles, pattern)
        self.treeDataWidget.completeTree(self.treeData)

    def getUserTitlesData(self):
        userTitles = []
        for idx in range(self.listData.count()):
            item = self.listData.item(idx)
            itemWdg = self.listData.itemWidget(item)

            nameField = itemWdg.nameField.text()
            coord = itemWdg.columns_combobox.currentData()
            if coord is None:
                continue
            userTitles += [(nameField, coord)]
        return userTitles

    def getTitlesData(self, rowNumber):
        dataList = []
        empty = False
        emptyCounter = 0
        column = 0
        while empty is False:
            column += 1
            cell = self.sheet.cell(rowNumber, column)
            value = cell.value

            if value:
                dataList.append((str(value), cell.coordinate))
                emptyCounter = 0
            else:
                emptyCounter += 1
            if emptyCounter > 100:
                empty = True
        return dataList

    def collectTreeData(self, nameColumn, startRow, userInpTitles, pattern):
        treeData = []
        empty = False
        emptyCounter = 0
        row = startRow

        while empty is False:
            row += 1

            cell = self.sheet.cell(row, nameColumn)
            value = cell.value

            if value:
                value = str(value)
                match = re.search(pattern, value)
                if not match:
                    continue

                shotData = self.collectShotData(row, userInpTitles)
                episodName = match.group(1)
                treeData = self.collectEpisodData(treeData, episodName, shotData)
                emptyCounter = 0
            else:
                emptyCounter += 1
            if emptyCounter > 100:
                empty = True
        # print(treeData)
        return treeData

    def collectShotData(self, row, userInpTitles):
        data = dict()
        for title in userInpTitles:
            field, coord = title
            column = self.sheet[coord].column
            value = self.sheet.cell(row, column).value
            if value:
                data[field] = str(value)
        return data

    def collectEpisodData(self, data, episodName, shotName):
        # found = False
        for episod in data:
            if episodName == episod.get('name'):
                episod['children'] += [shotName]
                return data
        data += [{'name': episodName, 'children': [shotName]}]
        return data


class ItemListWidget(QWidget):
    def __init__(self, parent, nameField, readOnly):
        super(ItemListWidget, self).__init__(parent)

        self.lay_main = QHBoxLayout()
        self.setLayout(self.lay_main)
        self.nameField = QLineEdit(nameField)
        self.nameField.setReadOnly(readOnly)
        self.nameField.setFixedWidth(300)
        self.columns_combobox = QComboBox()
        self.columns_combobox.setFixedWidth(200)

        # self.columns_combobox.

        self.lay_main.addWidget(self.nameField)
        self.lay_main.addWidget(self.columns_combobox)
        self.lay_main.addStretch()

    def setColumnsList_comboBox(self, titlesData):
        self.clearCombobBoxWidgetList(self.columns_combobox)
        for title in titlesData:
            self.columns_combobox.addItem(*title)

    def clearCombobBoxWidgetList(self, comboBoxWidget):
        prjCount = comboBoxWidget.count()
        while prjCount > 0:
            comboBoxWidget.removeItem(0)
            prjCount -= 1

# if __name__ == "__main__":
#     app = QApplication()
#     wdg = SheetDataWidget()
#     wdg.show()
#     app.exec_()
